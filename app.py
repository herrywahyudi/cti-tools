import streamlit as st
import pandas as pd
import re, io, os, zipfile, glob
from datetime import datetime, timedelta
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from pypdf.generic import NameObject, BooleanObject

st.set_page_config(page_title="CTI Tools", page_icon="🚢", layout="wide")

# ── Shared helpers ────────────────────────────────────────────────────────────

def fix_phone(p):
    s = re.sub(r'\D', '', str(p).split('.')[0])
    if s.startswith('62'): return s
    if s.startswith('0'):  return '62' + s[1:]
    if s.startswith('8'):  return '62' + s
    return s

def norm(text):
    return str(text).strip().lower().replace('_',' ').replace('-',' ').replace('/',' ')

def find_col(df, names):
    nc = {norm(c): c for c in df.columns}
    for n in names:
        if norm(n) in nc:
            return nc[norm(n)]
    return None

def clean_cell(v):
    v = str(v).strip()
    return '' if v.lower() == 'nan' else v

def clean_id(v):
    v = clean_cell(v)
    if not v: return ''
    try:
        n = float(v)
        if n.is_integer(): return str(int(n))
    except: pass
    return v[:-2] if v.endswith('.0') else v

def safe_filename(name):
    name = clean_cell(name) or 'candidate'
    for c in r'\/:*?"<>|': name = name.replace(c, '_')
    return ' '.join(name.split())

def parse_doc_text(docs):
    raw = str(docs).upper()
    for sep in [',',';','|','\n','\r','\t','(',')','{','}','[',']']:
        raw = raw.replace(sep, ' ')
    clean = ' '.join(raw.split())
    compact = clean.replace(' ','').replace('/','').replace('-','')
    return clean, compact

def has_doc(docs, keys):
    clean, compact = parse_doc_text(docs)
    tokens = set(clean.replace('/',' ').replace('-',' ').split())
    for key in keys:
        k = str(key).upper().strip()
        kc = k.replace(' ','').replace('/','').replace('-','')
        if k in clean or k in tokens or (kc and kc in compact):
            return True
    return False

def parse_service_date(row, sod_col):
    sod_raw = clean_cell(row.get(sod_col,'')) if sod_col else ''
    for fmt in ['%m/%d/%Y','%d/%m/%Y','%Y-%m-%d','%m-%d-%Y','%d-%m-%Y','%d %b %Y','%d %B %Y']:
        try:
            return (datetime.strptime(sod_raw, fmt) - timedelta(days=7)).strftime('%Y-%m-%d')
        except: pass
    return (datetime.today() - timedelta(days=7)).strftime('%Y-%m-%d')

# ── Seafarer Lookup logic ─────────────────────────────────────────────────────

LOOKUP_CONFIG = [
    {'label':'BST Baru',      'sheet':'New STCW',                'match':'phone','key':'Phone Number','filter_col':'Document Type',       'filter_val':'BST Baru',  'mode':'exact',    'date_col':None,        'date_filter':None},
    {'label':'Vaccine',       'sheet':'New Registration',         'match':'phone','key':'Phone Number','filter_col':None,                  'filter_val':None,        'mode':None,       'date_col':None,        'date_filter':None},
    {'label':'Passport',      'sheet':'New Passport',             'match':'phone','key':'Phone Number','filter_col':None,                  'filter_val':None,        'mode':None,       'date_col':None,        'date_filter':None},
    {'label':"Seaman's Book", 'sheet':'New Seamans Book',         'match':'phone','key':'Phone Number','filter_col':'Payment Status',      'filter_val':'Completed', 'mode':'exact',    'date_col':None,        'date_filter':None},
    {'label':'Medical',       'sheet':'Medical SL Request',       'match':'crew', 'key':'Crew ID Number','filter_col':None,               'filter_val':None,        'mode':None,       'date_col':None,        'date_filter':None},
    {'label':'MCV',           'sheet':'ATV and MCV Registration', 'match':'phone','key':'Phone Number','filter_col':'Document Type',       'filter_val':'MCV',       'mode':'contains', 'date_col':None,        'date_filter':None},
    {'label':'ATV',           'sheet':'ATV and MCV Registration', 'match':'phone','key':'Phone Number','filter_col':'Document Type',       'filter_val':'ATV',       'mode':'contains', 'date_col':None,        'date_filter':None},
    {'label':'Visa C1/D',     'sheet':'VISA APPLICATIONS',        'match':'name', 'key':'Name',        'filter_col':'Please select the type of visa you want to process','filter_val':'C1/D Visa','mode':'exact','date_col':'Added Time','date_filter':'current_year'},
    {'label':'Visa Schengen', 'sheet':'VISA APPLICATIONS',        'match':'name', 'key':'Name',        'filter_col':'Please select the type of visa you want to process','filter_val':'Schengen', 'mode':'contains','date_col':None,   'date_filter':None},
]

FILE_SHEETS = {
    'STCW':    ['New STCW'],
    'Vaccine': ['New Registration'],
    'Passport':['New Passport'],
    'Seaman':  ['New Seamans Book'],
    'Medical': ['Medical SL Request'],
    'ATV':     ['ATV and MCV Registration'],
    'Visa':    ['VISA APPLICATIONS'],
}

def build_match_set(df, cfg):
    key_col = cfg['key']
    if key_col not in df.columns:
        return set()
    filt = df
    if cfg['filter_col'] and cfg['filter_val'] and cfg['filter_col'] in df.columns:
        if cfg['mode'] == 'contains':
            filt = df[df[cfg['filter_col']].str.contains(cfg['filter_val'], na=False)]
        else:
            filt = df[df[cfg['filter_col']].str.strip() == cfg['filter_val'].strip()]
    if cfg['date_col'] and cfg['date_filter'] and cfg['date_col'] in filt.columns:
        filt = filt.copy()
        filt[cfg['date_col']] = pd.to_datetime(filt[cfg['date_col']], errors='coerce')
        if cfg['date_filter'] == 'current_year':
            filt = filt[filt[cfg['date_col']].dt.year == datetime.now().year]
    if cfg['match'] == 'phone':
        return set(filt[key_col].apply(fix_phone).dropna())
    elif cfg['match'] == 'crew':
        return set(filt[key_col].astype(str).dropna())
    elif cfg['match'] == 'name':
        return set(filt[key_col].str.lower().str.strip().dropna())
    return set()

def run_lookup(proc_file, lookup_files):
    proc = pd.read_excel(proc_file, header=0)
    proc = proc.iloc[1:].reset_index(drop=True)

    phone_col = find_col(proc, ['Phone Number ','Phone Number','Phone'])
    crew_col  = find_col(proc, ['Crew ID','CrewID','Crew Id'])
    fn_col    = find_col(proc, ['First Name ','First Name','FirstName'])
    ln_col    = find_col(proc, ['Unnamed: 12','Last Name','LastName','Surname'])
    tgt_col   = find_col(proc, ['Documents Processed by Seafarer','Documents Processed'])

    if not tgt_col:
        return None, 'Column "Documents Processed by Seafarer" not found in processing file.'

    proc['_phone'] = proc[phone_col].apply(fix_phone) if phone_col else ''
    proc['_crew']  = proc[crew_col].apply(lambda x: str(int(x)) if pd.notna(x) and str(x)!='nan' else '') if crew_col else ''
    if fn_col and ln_col:
        proc['_name'] = (proc[fn_col].fillna('') + ' ' + proc[ln_col].fillna('')).str.strip().str.lower()
    elif fn_col:
        proc['_name'] = proc[fn_col].fillna('').str.strip().str.lower()
    else:
        proc['_name'] = ''

    # Load lookup sheets from uploaded files
    loaded = {}  # sheet_name -> df
    for f in lookup_files:
        fname = f.name.upper()
        for key, sheets in FILE_SHEETS.items():
            if key.upper() in fname:
                for sheet in sheets:
                    try:
                        df = pd.read_excel(f, sheet_name=sheet)
                        loaded[sheet] = df
                        f.seek(0)
                    except: pass

    lookup_sets = {}
    log_lines = []
    for cfg in LOOKUP_CONFIG:
        sheet = cfg['sheet']
        if sheet not in loaded:
            log_lines.append(f"!! {cfg['label']}: sheet '{sheet}' not found")
            lookup_sets[cfg['label']] = set()
            continue
        s = build_match_set(loaded[sheet].copy(), cfg)
        lookup_sets[cfg['label']] = s
        fi = f" (filter: {cfg['filter_val']})" if cfg['filter_val'] else ''
        di = f" + {cfg['date_filter']}" if cfg['date_filter'] else ''
        log_lines.append(f"OK  {cfg['label']}: {len(s)} records{fi}{di}")

    def get_docs(row):
        docs = []
        for cfg in LOOKUP_CONFIG:
            s = lookup_sets.get(cfg['label'], set())
            if not s: continue
            key = row['_phone'] if cfg['match']=='phone' else row['_crew'] if cfg['match']=='crew' else row['_name']
            if key in s:
                docs.append(f"{cfg['label']} (done)")
        docs.append('GL (done)')
        return ', '.join(docs) if docs else ''

    proc['_result'] = proc.apply(get_docs, axis=1)
    filled = (proc['_result'] != '').sum()
    log_lines.append(f"\nMatched: {filled} of {len(proc)} rows")

    # Write back to Excel in memory
    proc_file.seek(0)
    wb = load_workbook(proc_file)
    ws = wb.active
    doc_col_idx = None
    for cell in ws[1]:
        if cell.value and 'Documents Processed' in str(cell.value):
            doc_col_idx = cell.column
            break
    if not doc_col_idx:
        return None, 'Column "Documents Processed by Seafarer" not found in Excel header.'
    for i, val in enumerate(proc['_result']):
        ws.cell(row=i+3, column=doc_col_idx).value = val if val else None

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, '\n'.join(log_lines)

# ── Service Agreement logic ───────────────────────────────────────────────────

CHECKBOX_MAPPING = [
    (['POLICE','POLICE CLEARANCE','SKCK'],                   'Check Box 1'),
    (['MARLINS','MARLIN'],                                    'Check Box 2'),
    (['PP','PASSPORT','E-PASSPORT','EPASSPORT'],             'Check Box 3'),
    (['SB','SEAMAN','SEAMAN BOOK'],                          'Check Box 4'),
    (['MCU','MEDICAL','MEDICAL CHECK'],                      'Check Box 5'),
    (['BST','STCW','BASIC SAFETY'],                         'Check Box 6'),
    (['ATV','AUSTRALIAN TRANSIT VISA','AUS TRANSIT'],        'Check Box 7'),
    (['MCV','MARITIME CREW VISA','AUSTRALIAN MARITIME'],     'Check Box 8'),
    (['C1D','C1/D','C1 D','US VISA','USA VISA'],             'Check Box 9'),
    (['RPX','EMBASSY SERVICE'],                              'Check Box 10'),
    (['SCH','SCHENGEN','SCHENGEN VISA'],                     'Check Box 11'),
    (['SCH PRIME','PRIME TIME','SCHENGEN PRIME'],            'Check Box 12'),
    (['MMR'],                                                'Check Box 13'),
    (['YF','YELLOW FEVER','YELLOW'],                        'Check Box 14'),
    (['GL','GUARANTEE LETTER'],                              'Check Box 15'),
    (['ADD','ADDITIONAL','PASSPORT SUPPORTING LETTER'],      'Check Box 16'),
]

def set_checkbox_direct(writer, field_name):
    yes = NameObject('/Yes')
    found = False
    for page in writer.pages:
        annots = page.get('/Annots')
        if not annots: continue
        for annot_ref in annots:
            annot = annot_ref.get_object()
            name = annot.get('/T')
            parent = annot.get('/Parent')
            parent_obj = parent.get_object() if parent else None
            parent_name = parent_obj.get('/T') if parent_obj else None
            check_name = str(name) if name else (str(parent_name) if parent_name else '')
            if check_name == field_name:
                found = True
                annot.update({NameObject('/AS'): yes})
                if parent_obj:
                    parent_obj.update({NameObject('/V'): yes, NameObject('/AS'): yes})
                else:
                    annot.update({NameObject('/V'): yes})
    return found

def fill_text_fields(writer, fields):
    for page in writer.pages:
        try:
            writer.update_page_form_field_values(page, fields, auto_regenerate=True)
        except TypeError:
            writer.update_page_form_field_values(page, fields)

def run_service_agreement(df, pdf_template_bytes):
    fn_col       = find_col(df, ['First Name ','First Name','FirstName'])
    ln_col       = find_col(df, ['Unnamed: 12','Last Name','LastName','Surname'])
    sod_col      = find_col(df, ['Sign On Date','SignOnDate','SOD','Embarkation Date'])
    position_col = find_col(df, ['Position Hired','Position','Title','Rank','Jabatan'])
    principal_col= find_col(df, ['Cruise Line ','Cruise Line','Principal','Line','Company'])
    crew_id_col  = find_col(df, ['Crew ID','CrewID','Crew Id','ID'])
    vessel_col   = find_col(df, ['Joining Ship /Vessel','Joining Ship','Ship','Vessel','Vessel Name','Kapal'])
    doc_col      = find_col(df, ['Documents Processed by Seafarer','Documents Processed'])

    if not doc_col:
        return None, 'Column "Documents Processed by Seafarer" not found.'

    zip_buffer = io.BytesIO()
    log_lines = []
    created = 0
    skipped = 0

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for idx, row in df.iterrows():
            docs = clean_cell(row.get(doc_col, ''))
            if not docs:
                skipped += 1
                continue

            # Build full name from first + last name columns
            first = clean_cell(row.get(fn_col, '')) if fn_col else ''
            last  = clean_cell(row.get(ln_col, '')) if ln_col else ''
            raw_name = (first + ' ' + last).strip() or f'candidate_{idx+1}'
            clean_name = safe_filename(raw_name)
            service_date = parse_service_date(row, sod_col)
            out_filename = f'Service Agreement_{clean_name}_{service_date}.pdf'

            reader = PdfReader(io.BytesIO(pdf_template_bytes))
            writer = PdfWriter()
            writer.clone_reader_document_root(reader)
            try: writer.set_need_appearances_writer()
            except: pass

            fields = {
                'Crew Name':      raw_name,
                'SEAFARER NAME':  raw_name,
                'SEAFARER NAME ': raw_name,
                'Date 1':         service_date,
                'DATE':           service_date,
                'CTI GROUP OFFICER': 'Puput Putri',
                'DATE_2':         service_date,
                'POSITION TITLE': 'Accounting Officer',
            }
            if position_col:
                v = clean_cell(row.get(position_col,''))
                fields.update({'Position':v, 'POSITION':v, 'POSITION / TITLE':v})
            if principal_col:
                v = clean_cell(row.get(principal_col,''))
                fields.update({'Principal Line':v, 'PRINCIPAL / LINE':v})
            if crew_id_col:
                v = clean_id(row.get(crew_id_col,''))
                fields.update({'Crew ID':v, 'SEAFARER ID':v})
            if vessel_col:
                v = clean_cell(row.get(vessel_col,''))
                fields.update({'Vessel Name':v, 'VESSEL NAME':v})

            fill_text_fields(writer, fields)

            ticked = []
            for keys, box_name in CHECKBOX_MAPPING:
                if has_doc(docs, keys):
                    set_checkbox_direct(writer, box_name)
                    ticked.append(box_name)

            try:
                if '/AcroForm' in writer._root_object:
                    writer._root_object['/AcroForm'].update(
                        {NameObject('/NeedAppearances'): BooleanObject(True)}
                    )
            except: pass

            pdf_out = io.BytesIO()
            writer.write(pdf_out)
            pdf_out.seek(0)
            zf.writestr(out_filename, pdf_out.read())
            created += 1
            log_lines.append(f"OK  {out_filename}  [{', '.join(ticked) or 'no checkboxes'}]")

    log_lines.append(f"\nCreated: {created} PDFs | Skipped (no docs): {skipped}")
    zip_buffer.seek(0)
    return zip_buffer, '\n'.join(log_lines)

# ── UI ────────────────────────────────────────────────────────────────────────

st.title('🚢 CTI Seafarer Tools')
st.caption('Internal tool for CTI Indonesia — Bali Office')

tab1, tab2 = st.tabs(['📋 Seafarer Lookup', '📄 Service Agreement'])

# ── TAB 1: Seafarer Lookup ────────────────────────────────────────────────────
with tab1:
    st.header('Seafarer Document Lookup')
    st.markdown('Fills the **Documents Processed by Seafarer** column in your Processing file by matching against all lookup logs.')

    col1, col2 = st.columns([1,1])
    with col1:
        st.subheader('1. Processing File')
        proc_file = st.file_uploader('Upload Processing.xlsx', type=['xlsx','xls'], key='proc')

    with col2:
        st.subheader('2. Lookup Files')
        lookup_files = st.file_uploader(
            'Upload all lookup files (STCW, Vaccine, Passport, Seaman\'s Book, Medical, ATV/MCV, Visa)',
            type=['xlsx','xls'], accept_multiple_files=True, key='lookup'
        )

    if proc_file and lookup_files:
        st.info(f'Processing file: **{proc_file.name}** | Lookup files: **{len(lookup_files)}** uploaded')
        if st.button('▶ Run Lookup', type='primary', key='run_lookup'):
            with st.spinner('Matching documents...'):
                result, log = run_lookup(proc_file, lookup_files)
            if result:
                st.success('Done!')
                st.code(log)
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                fname = proc_file.name.replace('.xlsx','') + f'_updated_{ts}.xlsx'
                st.download_button('⬇ Download Updated Excel', result, file_name=fname,
                                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            else:
                st.error(log)
    else:
        st.info('Upload your Processing file and all lookup files to get started.')

    with st.expander('ℹ Lookup rules'):
        st.markdown('''
| Document | Sheet | Match by | Condition |
|---|---|---|---|
| BST Baru | New STCW | Phone | Document Type = "BST Baru" |
| Vaccine | New Registration | Phone | Any record |
| Passport | New Passport | Phone | Any record |
| Seaman\'s Book | New Seamans Book | Phone | Payment Status = Completed |
| Medical | Medical SL Request | Crew ID | Any record |
| MCV | ATV and MCV Registration | Phone | Document Type contains MCV |
| ATV | ATV and MCV Registration | Phone | Document Type contains ATV |
| Visa C1/D | VISA APPLICATIONS | Name | Type = C1/D Visa + added this year |
| Visa Schengen | VISA APPLICATIONS | Name | Type contains Schengen |
''')

# ── TAB 2: Service Agreement ──────────────────────────────────────────────────
with tab2:
    st.header('Service Agreement Auto Fill')
    st.markdown('Generates filled PDF service agreements for each seafarer based on their documents.')

    col1, col2 = st.columns([1,1])
    with col1:
        st.subheader('1. Seafarer List')
        sa_data = st.file_uploader('Upload Excel (with Documents Processed column)', type=['xlsx','xls','csv'], key='sa_data')

    with col2:
        st.subheader('2. PDF Template')
        sa_pdf = st.file_uploader('Upload Service Agreement PDF template', type=['pdf'], key='sa_pdf')

    if sa_data and sa_pdf:
        st.info(f'Data: **{sa_data.name}** | Template: **{sa_pdf.name}**')
        if st.button('▶ Run Auto Fill', type='primary', key='run_sa'):
            with st.spinner('Generating PDFs...'):
                if sa_data.name.endswith('.csv'):
                    df = pd.read_csv(sa_data)
                else:
                    df = pd.read_excel(sa_data)
                pdf_bytes = sa_pdf.read()
                zip_out, log = run_service_agreement(df, pdf_bytes)

            if zip_out:
                st.success('Done!')
                st.code(log)
                ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                st.download_button('⬇ Download All PDFs (ZIP)', zip_out,
                                   file_name=f'Service_Agreements_{ts}.zip',
                                   mime='application/zip')
            else:
                st.error(log)
    else:
        st.info('Upload your seafarer list and the PDF template to get started.')

    with st.expander('ℹ Supported document keywords'):
        st.markdown('''
| Keyword in cell | Checkbox ticked |
|---|---|
| SKCK / Police | Police Clearance |
| Marlins | Marlins Test |
| PP / Passport | Passport |
| SB / Seaman | Seaman\'s Book |
| MCU / Medical | Medical Check |
| BST / STCW | Basic Safety Training |
| ATV | Australian Transit Visa |
| MCV | Maritime Crew Visa |
| C1/D / C1D | US Visa C1/D |
| SCH / Schengen | Schengen Visa |
| MMR | MMR Vaccine |
| YF / Yellow Fever | Yellow Fever |
| GL | Guarantee Letter |
| Additional | Additional Doc |
''')

st.divider()
st.caption('CTI Indonesia · Bali · Internal Use Only')
